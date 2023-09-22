# import the necessary packages
from read_cols import read_both
import numpy as np
from functools import reduce
from itertools import combinations
from SpezRedBlackCodes import get_all_red_black_spez_codes

import sys
np.set_printoptions(threshold=sys.maxsize)

def print_bin_str(arr,length):
    print()
    for d in arr:
        print(np.binary_repr(d, width=length))
    print()
    print("size of input arr:", len(arr))
    print()
    
    
def binA(llst_nbrs):

    # evaluating shape of A
    lst_ind_lasts = [row[-1] for row in llst_nbrs]
    m = len(llst_nbrs)      # hight
    n = max(lst_ind_lasts)  # width

    A = np.zeros((m, n), dtype=int)

    # Fill 1's in A at row index of lst and columns index = lst at respective row and entry
    for k in range(m):
        for ind in llst_nbrs[k]:
            A[k, ind - 1] = 1  # A is bin matrix of all rows and columns including singletons

    return A
 
# functions to set domain D and splitting
D = [7, 11, 19, 35, 67, 131, 13, 21, 37, 69, 133, 25, 41, 73, 137, 49, 81, 145, 97, 161, 193, 14, 22, 38,
              70, 134, 26, 42, 74, 138, 50, 82, 146, 98, 162, 194, 28, 44, 76, 140, 52, 84, 148, 100, 164, 196, 56,
              88, 152, 104, 168, 200, 112, 176, 208, 224]
def concat_two_codes(codes2,codes1,bit_len_2,bit_len_1):
    shifted_codes2 = [c2 << bit_len_1 for c2 in codes2]
    codes2_con_codes1 = [sh_c2 + c1 for sh_c2 in shifted_codes2 for c1 in codes1]
    return codes2_con_codes1
def extend_exclusively_two_codes(codes2,codes1,bit_len_2,bit_len_1):   
    #shift code2 to the left by bit_len_1, e.g., 00101001 -> 0010100100000000
    shifted_codes2 = [c2 << bit_len_1 for c2 in codes2]   
    #fill right with all 1's to be "unblocking" in code2
    num_right = 2**bit_len_1 - 1
    excl_codes2 = [sh_c2 + num_right for sh_c2 in shifted_codes2]    
    #fill left with all 1's to be "unblockin" in code2
    num_left = 2**(bit_len_1 + bit_len_2) - (num_right + 1)
    excl_codes1 = [num_left + c1 for c1 in codes1]    
    return excl_codes1 + excl_codes2
def extend_exclusively_two_codes_7(codes2,codes1,bit_len_2,bit_len_1,bit_pos_7_red,bit_pos_7_black):   
    #shift code2 to the left by bit_len_1, e.g., 00101001 -> 0010100100000000
    shifted_codes2 = [c2 << bit_len_1 for c2 in codes2]   
    #fill right with all 1's to be "unblocking" in code2
    num_right = 2**bit_len_1 - 1  - 2**bit_pos_7_black
    excl_codes2 = [sh_c2 + num_right for sh_c2 in shifted_codes2]    
    #fill left with all 1's to be "unblockin" in code2
    num_right = 2**bit_len_1 - 1
    num_left = 2**(bit_len_1 + bit_len_2) - (num_right + 1)  - 2**(bit_pos_7_red + bit_len_1)
    excl_codes1 = [num_left + c1 for c1 in codes1]    
    return excl_codes1 + excl_codes2
def get_H(le, nbr_0s):
    # Get all possible combinations of positions for the unset bits
    unset_bits_positions = list(combinations(range(le), nbr_0s))
    H = []
    # Set all bits to 1
    num = (1 << le) - 1
    for positions in unset_bits_positions:
        # Unset the bits at the specified positions
        temp_num = num
        for pos in positions:
            temp_num &= ~(1 << pos)
        H.append(temp_num)
    return H
def get_H(all_r, le):
    # Extract the left and right four bits from all_r
    left_h = (all_r >> 22) & 0b1111
    right_h = (all_r >> 18) & 0b1111

    # Get all possible combinations of positions for the unset bits
    def get_combinations(h, le):
        unset_bits_positions = [i for i in range(4) if ((h >> i) & 1) == 0]
        return list(combinations(unset_bits_positions, le))

    # Generate all numbers that have precisely le bits unset among the positions where h has unset bits
    def generate_numbers(h, combinations):
        numbers = []
        for positions in combinations:
            num = 0b1111 #h
            for pos in positions:
                num &= ~(1 << pos)
            numbers.append(num)
        return numbers

    # Generate H_left and H_right
    H_left = [(h << 4) + 15 for h in generate_numbers(left_h, get_combinations(left_h, le))]
    H_right = [h + 0b11110000 for h in generate_numbers(right_h, get_combinations(right_h, le))]          #+240

    return H_left, H_right

def split_bin(B, b_le_right):
    r_1s = 2**b_le_right - 1
    B_right = [b & r_1s for b in B]
    B_left  = [b >> b_le_right for b in B]
    return B_left, B_right

def get_7_poss(code):
    h, s1, s2 = code
    pos = s1.index('7')
    if pos % 4 == 0: #black
        pos_7_black = pos // 4
        pos_7_red = s2.index('7') // 4 + 4
    else:
        pos_7_red = (pos - 2) // 4
        pos_7_black = (s2.index('7') - 2) // 4 + 5

    bit_pos_7_black = 8 - pos_7_black
    bit_pos_7_red = 8 - pos_7_red

    return bit_pos_7_red, bit_pos_7_black

def insert_zero_bit(D, bit_pos_7):
    if bit_pos_7 == 0:
        D_7 = [d << 1 for d in D]
    elif bit_pos_7 == 8:
        D_7 = D
    else:
        left = 2 ** 8 - 2 ** bit_pos_7
        righ = 2 ** bit_pos_7 - 1
        D_7 = [((d & left) << 1) + (d & righ) for d in D]
    return D_7
    
    
    
# use codes assigned for normal, or zentral, respectively    
def modify_bits(c_pass, all_r):
    # Define masks for the left and right four bit positions
    left_mask  = 0b11110000000000000000000000
    right_mask = 0b00001111000000000000000000

    # Check if all_r has any unset bit in the left and right four bit positions
    left = (all_r & left_mask) != left_mask
    right = (all_r & right_mask) != right_mask

    # If left is True, unset each bit in the left four bit positions in all_r,
    # if it is unset in c_pass but wasn't in all_r
    if left:
        all_r &= ~((~c_pass & all_r) & left_mask)

    # If right is True, unset each bit in the right four bit positions in all_r,
    # if it is unset in c_pass but wasn't in all_r
    if right:
        all_r &= ~((~c_pass & all_r) & right_mask)

    return all_r

# get data restricted to entries that have actual code and get original indices
def get_actual_code_data(codes): # but must be all keys actually !!!
    actual_codes = []
    original_indices = []
    values = {'2', '3'}
    for v, code in enumerate(codes):        
        if any(x in values for x in code[0]):
            actual_codes.append(code)
            original_indices.append(v)
    return actual_codes, original_indices

# get data of all keys, codes has to page end 60
def get_actual_code_data_of_all_keys(codes): # but must be all keys actually !!!
    actual_codes = []
    original_indices = []
    for v, code in enumerate(codes):        
        if any(x == '7' for x in code[1]):
            actual_codes.append(code)
            original_indices.append(v)
    return actual_codes, original_indices

# extract bin codes for Zylinders from read data
def get_R_pr(codes):
    codes, original_indices = get_actual_code_data(codes)
    values = {'2', '3', '4', '7'}
    R_pr = []
    ind_print = 0
    for h, s1, s2 in codes:
        ind_print += 1
        p_h = [k for k, x in enumerate(h) if x in values]
        p_s1 = [k for k, x in enumerate(s1) if x in values]
        p_s2 = [k for k, x in enumerate(s2) if x in values]
        if p_h[0] % 4 == 0: #black
            int_h = 15 - sum([2 ** (3 - p // 4) for p in p_h])
            int_h = (15 << 4) + int_h
        else:
            int_h = 15 - sum([2 ** (3 - (p - 2) // 4) for p in p_h])
            int_h = (int_h << 4) + 15
        if p_s1[0] % 4 == 0: #black
            int_s1 = 31 - sum([2 ** (4 - p // 4) for p in p_s1])
            int_s2 = 15 - sum([2 ** (3 - (p - 2) // 4) for p in p_s2])
            r = (int_h << 18) + ((2 ** 9 - 1) << 9) + (int_s1 << 4) + int_s2
        else:
            int_s1 = 15 - sum([2 ** (3 - (p - 2) // 4) for p in p_s1])
            int_s2 = 31 - sum([2 ** (4 - p // 4) for p in p_s2])   
            r = (int_h << 18) + (int_s1 << 14) + (int_s2 << 9) + (2 ** 9 - 1)
        R_pr.append(r)
    return R_pr, original_indices

# extract bin codes for Schlüssel- and Zentralcodes from read data
def get_C_pr(codes):
    #codes, original_indices = get_actual_code_data(codes)
    codes, original_indices = get_actual_code_data_of_all_keys(codes)
    values={'2', '3', '4', '7'}
    C_pr = []
    for h, s1, s2 in codes:
        p_h = [k for k, x in enumerate(h) if x in values]
        p_s1 = [k for k, x in enumerate(s1) if x in values]
        p_s2 = [k for k, x in enumerate(s2) if x in values]
        p_h_black = [k for k, x in enumerate(h) if x in values and k % 4 == 0]
        p_h_red = [k for k, x in enumerate(h) if x in values and k % 4 != 0]
        p_s1_black = [k for k, x in enumerate(s1) if x in values and k % 4 == 0]
        p_s2_black = [k for k, x in enumerate(s2) if x in values and k % 4 != 0]
        p_s1_red = [k for k, x in enumerate(s1) if x in values and k % 4 != 0]
        p_s2_red = [k for k, x in enumerate(s2) if x in values and k % 4 == 0]
        
        int_h_black = 15 - sum([2 ** (3 - p // 4) for p in p_h_black])
        int_h_black = (15 << 4) + int_h_black
        int_s1_black = 31 - sum([2 ** (4 - p // 4) for p in p_s1_black])
        int_s2_black = 15 - sum([2 ** (3 - (p - 2) // 4) for p in p_s2_black])
        r_black = (int_h_black << 18) + ((2 ** 9 - 1) << 9) + (int_s1_black << 4) + int_s2_black    
        
        int_h_red = 15 - sum([2 ** (3 - (p - 2) // 4) for p in p_h_red])
        int_h_red = (int_h_red << 4) + 15        
        int_s1_red = 15 - sum([2 ** (3 - (p - 2) // 4) for p in p_s1_red])
        int_s2_red = 31 - sum([2 ** (4 - p // 4) for p in p_s2_red])   
        r_red = (int_h_red << 18) + (int_s1_red << 14) + (int_s2_red << 9) + (2 ** 9 - 1)      
        
        C_pr.append(r_red & r_black)        
    return C_pr, original_indices

        
# extract R_pr, C_pr from reading and get whole final_data for latter writting
def get_prev_data():
    final_data = read_both()
    R_pr, R_original_indices = get_R_pr(final_data["normalzylinder"][1])
    C_pr, C_original_indices = get_C_pr(final_data["schlussel"][1])     #currently all stored till page end        
    return R_pr, C_pr, R_original_indices, C_original_indices, final_data


# extract entire code bit number used in normal only (same if for zentral)
def get_domain_int(R_pr, C_pr, D, final_data, mixed=False):
    all_r = reduce(lambda x, y: x & y, R_pr)
    res_HK = True                                 # dummy
    if res_HK:
        all_r = modify_bits(C_pr[0], all_r)       # c_pass is C_pr[0] for now.
    
    mask_left = 0b00000000111111111000000000
    mask_righ = 0b00000000000000000111111111
    
    bit_pos_7_red, bit_pos_7_black = get_7_poss(final_data["schlussel"][1][0])
    
    if (all_r & mask_left) == 0:
        D_left = D
        D_left = insert_zero_bit(D, bit_pos_7_red)
    else:
        D_left = [511]
    if (all_r & mask_righ) == 0:
        D_righ = D
        D_righ = insert_zero_bit(D, bit_pos_7_black)
    else:
        D_righ = [511]
    
    D_ext = extend_exclusively_two_codes(D_left, D_righ, 9, 9)
    if mixed:
        D_spe = get_all_red_black_spez_codes(bit_pos_7_red, bit_pos_7_black)                         # trying
        D_ext = D_ext + D_spe[0:-1]
    #D_ext = extend_exclusively_two_codes_7(D_left, D_righ, 9, 9, bit_pos_7_red, bit_pos_7_black)
    H_left, H_righ = get_H(all_r, 2)  
    H = H_left + H_righ
    D_ext = concat_two_codes(H, D_ext, 8, 18)
    
    return D_ext
    

"""
# getting appropriate datas of current R_pr, C_pr, domain D_ext and A
R_pr, C_pr, R_original_indices, C_original_indices, final_data = get_prev_data()

resulting_middle_N = [[int(num) for num in string.split()] for string in final_data["normal_middle"]]
A = binA(resulting_middle_N)

D_ext = get_domain_int(R_pr, C_pr, D) #this D_ext is domain ready for computing
"""

# setting up functions for solving
# first, we need extended matrix 
def get_A_ex(A, Inz_ex):  # User input, e.g. Inz_ex = [{9, 10}, {10}, {3, 10}], these are col-indices 1-indexed
    m, n = A.shape
    m_add = len(Inz_ex)
    n_add = max(max(s) for s in Inz_ex)  # find the maximum index in Inz_ex
    m_ex = m + m_add
    n_ex = max(n, n_add)
    A_ex = np.zeros((m_ex, n_ex), dtype=int)
    A_ex[:m, :n] = A
    # Construct A_add from Inz_ex
    A_add = np.zeros((m_add, n_add), dtype=int)
    for i in range(m_add):
        for idx in Inz_ex[i]:
            A_add[i, idx-1] = 1  # subtract 1 because indices in Inz_ex are 1-indexed
    A_ex[m:, :n_add] = A_add
    return A_ex

def get_A_ex(A, Inz_ex, InzC_ex={}): # User input, e.g. Inz_ex = [{9, 10}, {10}, {3, 10}], InzC_ex = {1: {6}, 2: {6, 7}}
    m, n = A.shape
    m_add = len(Inz_ex)
    Inz_ex = [frozenset(s) for s in Inz_ex]
    n_add = max(max(s) for s in set(Inz_ex).union(map(frozenset, InzC_ex.values()))) if Inz_ex or InzC_ex.values() else 0
    # find the maximum index in Inz_ex and InzC_ex
    m_ex = m + m_add
    n_ex = max(n, n_add)
    A_ex = np.zeros((m_ex, n_ex), dtype=int)
    A_ex[:m, :n] = A
    for i in range(m_add):
        for idx in Inz_ex[i]:
            A_ex[m + i, idx-1] = 1
    for k, val in InzC_ex.items():
        for j in val:
            #A_ex[k-1, j-1] = 1     # A_ex[k-1, abs(j)-1] = (-1 < j)*1                                                    ########
            A_ex[k-1, abs(j)-1] = (-1 < j)*1
            print(j)
    print(A_ex[6])
    return A_ex


from z3 import *

def CSP_ex(A, A_ex, D, R_pr, C_pr, repl, bitlength=17): # repl are those range(m)-indices that get replaced
    defa = 2 ** bitlength - 1
    m, n = A.shape
    m_ex, n_ex = A_ex.shape
    print("m_ex, n_ex=", m_ex, n_ex)
    opt = Optimize()
    #opt.set("model.completion", True)
    opt.set("timeout", 88888)

    R_ex = [BitVec(f"rs_{i + 1}", bitlength) for i in range(m_ex)]
    C_ex = [BitVec(f"cs_{j + 1}", bitlength) for j in range(n_ex)]

    opt.add([Or([r == d for d in D]) for r in R_ex])
        
    opt.add([R_ex[i] & C_ex[j] == C_ex[j] for i in range(m_ex) for j in range(n_ex) if A_ex[i, j] == 1])
    opt.add([R_ex[i] & C_ex[j] != C_ex[j] for i in range(m_ex) for j in range(n_ex) if A_ex[i, j] == 0])

    # Minimize the difference between R_pr and R_ex and between C_pr and C_ex
    # Exclude elements of R_pr corresponding to the first index in each tuple of repl from the minimization objective
    excluded_indices = [t[0] for t in repl]

    # Verwenden Sie die If-Funktion, um BoolRef-Objekte in Zahlen umzuwandeln
    sum_R = sum([If(R_pr[i] != R_ex[i], 1, 0) for i in range(m) if i not in excluded_indices])
    sum_C = sum([If(C_pr[j] != C_ex[j], 1, 0) for j in range(n)])
    opt.minimize(sum_R + sum_C)
    #opt.minimize(sum([R_pr[i] != R_ex[i] for i in range(m) if i not in excluded_indices]) + sum([C_pr[j] != C_ex[j] for j in range(n)]))

    if opt.check() == sat:
        model = opt.model()
        R_ex_sol = [model.evaluate(R_ex[i]).as_long() for i in range(m_ex)]
        C_ex_sol = [model.evaluate(C_ex[j]).as_long() for j in range(n_ex)]
        viol_R = [i for i in range(m) if R_pr[i] != R_ex_sol[i]]
        viol_C = [j for j in range(n) if C_pr[j] != C_ex_sol[j]]
        print(f"Indices where R_pr and R_ex differ: {viol_R}")
        print(f"Indices where C_pr and C_ex differ: {viol_C}")
        return R_ex_sol, C_ex_sol, viol_R, viol_C
    else:
        return [], [], [], []

"""
# test execution ---------------------------
#Inz_ex, InzC_ex = get_Inz_ex_InzC_ex(n, str_Inz_ex, str_InzC_ex)              #str_Inz_ex, str_InzC_ex user input
# define extension
Inz_ex = [{1,3,58}]
Inz_ex = [{1,3,5,6}, {1,5,6,11}, {1,3,58}, {1,7,59}, {1,8,60}] # pass muss drüber


Inz_ex = [{1,58}, {1,59}, {1,60}, {1,3,61}, {1,5,62}, {1,5,63}]
InzC_ex = {1: {6, 58}, 2: {58}}

A_ex = get_A_ex(A, Inz_ex, InzC_ex)


repl = [(7, 8)] #zylinder not delivered yet -> can be replaced by new index 2nd tuple entry7 (1-indexed)
repl = []
# ------------------------------------------


R_ex_sol, C_ex_sol, viol_R, viol_C = CSP_ex(A, A_ex, D_ext, R_pr, C_pr, repl, bitlength=26)

print("checking sat:")
m_ex, n_ex = A_ex.shape
print(A_ex.shape)
print(all([R_ex_sol[i] & C_ex_sol[j] == C_ex_sol[j] for i in range(m_ex) for j in range(n_ex) if A_ex[i, j] == 1]))
print(all([R_ex_sol[i] & C_ex_sol[j] != C_ex_sol[j] for i in range(m_ex) for j in range(n_ex) if A_ex[i, j] == 0]))
"""


# getting result in writting format_____
def transform_to_codes(R_ex_sol, pass_entries):
    hk_pass, s1_pass, s2_pass = pass_entries
    codes = []
    for num in R_ex_sol:
        # Convert the number to binary and remove the '0b' prefix
        binary = bin(num)[2:]
        # Pad the binary number with zeros at the start to make it 26 bits long
        binary = binary.zfill(26)
        
        # Split binary into hk, s1, and s2 parts
        hk_left, hk_right = binary[:4], binary[4:8]
        s_left, s_right = binary[8:17], binary[17:]
        
        # Initialize hk, s1, and s2 lists
        hk = [''] * 18 + ['Z'] + ['']
        s1 = [''] * 3 + [None] + [''] * 3 + [None] + [''] * 3 + [None] + [''] * 3 + [None] + [''] * 3 + [None]
        s2 = list(s1)
        
        # Fill hk based on hk_left and hk_right
        for i, bit in enumerate(hk_right):
            if bit == '0':
                hk[i*4] = hk_pass[4*i] # '2'
        for i, bit in enumerate(hk_left):
            if bit == '0':
                hk[i*4+2] = hk_pass[i*4+2] # '2'
        
        # Fill s1 and s2 based on s_left and s_right
        s_right_s1, s_right_s2 = s_right[:5], s_right[5:]
        for i, bit in enumerate(s_right_s1):
            if bit == '0':
                s1[i*4] = s1_pass[i*4] # '2'
        for i, bit in enumerate(s_right_s2):
            if bit == '0':
                s2[i*4+2] = s2_pass[i*4+2] # '2'
        
        s_left_s1, s_left_s2 = s_left[:4], s_left[4:]
        for i, bit in enumerate(s_left_s1):
            if bit == '0':
                s1[i*4+2] = s1_pass[i*4+2] # '2'
        for i, bit in enumerate(s_left_s2):
            if bit == '0':
                s2[i*4] = s2_pass[i*4] # '2'
        
        # Add the three lists to the codes list
        codes.append([hk, s1, s2])
    
    return codes

def get_extension_ordered_codes(R_pr, R_ex_sol, viol_R):
    m, m_ex = len(R_pr), len(R_ex_sol)
    final_R = R_pr + [R_ex_sol[i] for i in viol_R] + R_ex_sol[m:m_ex]
    
    replaced_by = {viol_R[i]: m + i for i in range(len(viol_R))}   
    return final_R, replaced_by

def get_extension_middles(R_pr, C_pr, R_ex_sol, C_ex_sol, viol_R, viol_C):
    final_R, R_replaced_by = get_extension_ordered_codes(R_pr, R_ex_sol, viol_R)
    final_C, C_replaced_by = get_extension_ordered_codes(C_pr, C_ex_sol, viol_C)
    m_ex, n_ex = len(R_ex_sol), len(C_ex_sol)
    m_final, n_final = len(final_R), len(final_C)
    final_A = np.full((m_final, n_final), False, dtype=bool)
    for i in range(m_final):
        for j in range(n_final):
            final_A[i, j] = ((final_R[i] & final_C[j]) == final_C[j])
            
    #restrict new keys to their inc rows
    defa = 0b11111111111111111111111111
    final_C = C_pr + [reduce(lambda x,y: x & y, [final_R[i] for i in range(m_final) if final_A[i, j]], defa) for j in range(len(C_pr), n_final)]#range(n_final)]
    
    #print("final_C")
    #print_bin_str(final_C, 26)
    
    middles = []
    for a in final_A:
        middles.append(' '.join(map(str, [j + 1 for j in range(n_final) if a[j]])))
    return final_R, final_C, middles, R_replaced_by, C_replaced_by


        
    
    
#_______________________________________


#codes = transform_to_codes(R_ex_sol, final_data["schlussel"][1][0]) # testing


#final_R = get_extension_ordered_codes(R_pr, R_ex_sol, viol_R)
#final_C = get_extension_ordered_codes(C_pr, C_ex_sol, viol_C)

def impose_both_7_into_C(final_C, final_data):
    bit_pos_7_red, bit_pos_7_black = get_7_poss(final_data["schlussel"][1][0])
    both_fix = 0b11111111111111111111111111 - (2**(bit_pos_7_red + 9) + 2**bit_pos_7_black)
    final_C_with_both_7 = [c & both_fix for c in final_C]
    return final_C_with_both_7

"""
final_R, final_C, middles, R_replaced_by, C_replaced_by = get_extension_middles(R_pr, C_pr, R_ex_sol, C_ex_sol, viol_R, viol_C)

final_C = impose_both_7_into_C(final_C) #get 0 at both 7-pos
      

final_R_data = transform_to_codes(final_R, final_data["schlussel"][1][0])
final_C_data = transform_to_codes(final_C, final_data["schlussel"][1][0])

#updating final_data for writting
m_ex, n_ex = len(final_R), len(final_C)
final_m, final_n = len(final_R_data), len(final_C_data)
final_data["normalzylinder"] = ([str(i + 1) for i in range(final_m)], final_R_data)
final_data["schlussel"] = ([str(j + 1) for j in range(final_n)], final_C_data)   ###### !!!!!!!!!!!!
final_data["normal_middle"] = middles

print(f"len(final_C_data)={len(final_C_data)}, n_ex={n_ex}")
"""



# writting updated final_data -------------------------------------------------
import openpyxl
from openpyxl.styles import Font
from colors import read_colors, append_colors
from copy import deepcopy
                                               
# Create a new font object with the desired properties  
grey = "FF4B4B4B"       
green = "FF006400"
blue = "FF00008B"
                           
fontgrey = Font(color=grey, size=16)
fontgreen = Font(color=green, size=16, bold=True)    # FF006400   FF003300
fontblue = Font(color=blue, size=16, bold=True)
fontblue12 = Font(color=blue, size=12, bold=True)
fontgrey12 = Font(color=grey, size=12)
fontgreen12 = Font(color=green, size=12, bold=True)  # FF006400   FF003300



# This class refers to a single group of rows in the excel sheet
class MiddleObject:
    def __init__(self):
        self.triplets = []

    def add_triplet(self, triplet):
        self.triplets.append(triplet)

    def __repr__(self) -> str:
        for triplet in self.triplets:
            print(triplet)
        return ''
    
# This class refers to one particular row in the excel sheet
class Triplet:
    def __init__(self,row1,row2,row3) -> None:
        temp = ["" for _ in range(15)]
        if row1 is None:
            row1 = temp        
        if row2 is None:
            row2 = temp
        if row3 is None:
            row3 = temp

        self.row1 = row1
        self.row2 = row2
        self.row3 = row3
        
    
    def exists(self):
        return any(self.row1) or any(self.row2) or any(self.row3)
    
    def __repr__(self) -> str:
        print(*self.row1)
        print(*self.row2)
        print(*self.row3)
        return ''


wb = openpyxl.load_workbook("Data/raw.xlsx")
sheet_one = wb[wb.sheetnames[0]]
sheet_N = wb[wb.sheetnames[1]]
sheet_S = wb[wb.sheetnames[2]]
sheet_Z = wb[wb.sheetnames[3]]
sheet_ZC = wb[wb.sheetnames[4]]

"""
# all required data fetched by read_cols script in a dict   
#final_data = read_both()                                    
# all required data fetched by the colors script
colors_data = read_colors()
#colors_data['normalzylinder'] += colors_data['normalzylinder'][29:]     # temporarily fixed for 33 +  keys      
append_colors(colors_data, final_data, 'normalzylinder', final_m) 


wb = openpyxl.load_workbook("Data/raw.xlsx")
sheet_one = wb[wb.sheetnames[0]]
sheet_N = wb[wb.sheetnames[1]]
sheet_S = wb[wb.sheetnames[2]]
sheet_Z = wb[wb.sheetnames[3]]
sheet_ZC = wb[wb.sheetnames[4]]

# refers to page breaks for both sheets, -1 initially
zentral_page_break = -1
normal_page_break = -1

# used for assigning scores to each row, while sorting them, one for red and one for black
red_scores = {}
black_scores = {}

# used for spaces in zentralzylinder and normalzylinder
spaces = {}
"""

# assigns score to each unique first row
def return_score(data, dc):
    for triplet in data:
        row = triplet[0]
        score = 0
        for index, value in enumerate(row):
            if value in ["","Z","B"]:
                continue
            score += (2500 - 100 * index) + float(value)
        # print("added key", ",".join(row))
        dc[",".join(row)] = score



def check_default(a):
    for item in a:
        if item == 0:
            return False
    return True


# this function performs grouping for both zentralzylinder and normalzylinder
def group_left(name, middle, final_data, colors_data, spaces):
    global red_scores, black_scores
    red_scores = {}
    black_scores = {}

    # fetching current indices, data, middle values
    index_list = final_data[name][0]
    data_list = final_data[name][1]
    middle_data = final_data[middle]

    # seperate lists for indices, data, middle values
    index_black_codes = []
    index_red_codes = []
    data_black_codes = []
    data_red_codes = []
    middle_red_codes = []
    middle_black_codes = []

    # fetching color codes
    zentral_color_codes = colors_data[name]

    for index, color_code in enumerate(zentral_color_codes):
        # segregating into red and black based on S1,S2
        if 0 in color_code[3] or 0 in color_code[8]:
            index_black_codes.append(index_list[index])
            data_black_codes.append(data_list[index])
            middle_black_codes.append(middle_data[index])
        else:
            index_red_codes.append(index_list[index])
            data_red_codes.append(data_list[index])
            middle_red_codes.append(middle_data[index])

    # calculating page break
    page_break = len(index_red_codes)

    original_red_indices = deepcopy(index_red_codes)
    # calculating score for red and black, and sorting
    return_score(data_red_codes, red_scores)

    if index_red_codes:

        # sorting all 3 lists with red scores as the attribute
        index_red_codes, data_red_codes = zip(
            *(
                sorted(
                    zip(index_red_codes, data_red_codes),
                    key=lambda x: red_scores[",".join(x[1][0])],
                    reverse=True,
                )
            )
        )
        middle_red_codes = [
            middle_red_codes[original_red_indices.index(x)] for x in index_red_codes
        ]

    original_black_indices = deepcopy(index_black_codes)
    return_score(data_black_codes, black_scores)
    if index_black_codes:
        # sorting all 3 lists with black scores as the attribute
        index_black_codes, data_black_codes = zip(
            *(
                sorted(
                    zip(index_black_codes, data_black_codes),
                    key=lambda x: black_scores[",".join(x[1][0])],
                    reverse=True,
                )
            )
        )
        middle_black_codes = [
            middle_black_codes[original_black_indices.index(x)]
            for x in index_black_codes
        ]

    # joining lists
    index_red_codes, data_red_codes = (
        list(index_red_codes),
        list(data_red_codes),
    )
    index_black_codes, data_black_codes = (
        list(index_black_codes),
        list(data_black_codes),
    )

    # to add indices for which spaces are required within both groups
    spaces[name] = []

    # to add indices for which spaces are required within red group
    for index, triplet in enumerate(data_red_codes):
        if index == 0:
            continue
        if data_red_codes[index - 1][0] != triplet[0]:
            spaces[name].append(index)

    # to add indices for which spaces are required within black group
    for index, triplet in enumerate(data_black_codes):
        if index == 0:
            continue
        if data_black_codes[index - 1][0] != triplet[0]:
            spaces[name].append(index + len(data_red_codes))

    # combining lists
    index_red_codes.extend(index_black_codes)
    data_red_codes.extend(data_black_codes)
    middle_red_codes.extend(middle_black_codes)

    # reassigning to final data
    final_data[name] = (index_red_codes, data_red_codes)
    final_data[middle] = middle_red_codes

    # returning page break
    return page_break, spaces


# writes both zentralzylinder and normalzylinder
def write_left(name, sheet, page_break, middle_name, viol_R, R_replaced_by, final_data, spaces, m, dict_newR_I):

    # the required data is fetched from the dictionary
    index_list = final_data[name][0]
    data_list = final_data[name][1]  #zylinder-codes
    
    replacing_R = set(R_replaced_by.values())
    
    dict_newI_R = {str(i): su for su, i in dict_newR_I.items()}                ############################

    # starting coordinates
    x_cord = 4
    y_cord = 1
    
    # writing the data to excel
    for index, number in enumerate(index_list):
        
        x_cord_end = x_cord+3*len(final_data[middle_name][index].triplets)
        row_indices = [_ for _ in range(x_cord,x_cord_end+1) if (_%57==0 and _!=x_cord_end)]
        
        # checks if there is a page break at given x coordinate
        if index == page_break or row_indices:
            x_cord -= 1
            x_cord += 60 - x_cord % 57
        # checks if a blank row is required at that index
        elif index in spaces[name]:
            x_cord += 3

        
        # print(f"INDEX VALUE at {x_cord},{y_cord} is {number}")
        if x_cord%3 == 0:
            x_cord+=1
        
        if int(number) - 1 in viol_R:
            sheet.cell(row=x_cord, column=y_cord).font = fontgrey                # try grey      index
            sheet.cell(row=x_cord, column=y_cord).value = str(number) + ' →' + str(R_replaced_by[int(number) - 1] + 1)
            
        elif int(number) - 1 in replacing_R:
            sheet.cell(row=x_cord, column=y_cord).font = fontgreen #
            sheet.cell(row=x_cord, column=y_cord).value = number
            
        elif int(number) > m:
            sheet.cell(row=x_cord, column=y_cord).font = fontblue #
            sheet.cell(row=x_cord, column=y_cord).value = number + ' ' + dict_newI_R[number]         #####
            
        else:    
            sheet.cell(row=x_cord, column=y_cord).value = number

        for row_index, row in enumerate(data_list[index], x_cord):
            for value_index, value in enumerate(row, y_cord + 1):
                sheet.cell(row=row_index, column=value_index).value = value
        
        # print(len(final_data[middle_name][index].triplets))

        # offset refers to the row that has to be skipped if len of triplets is more than 1
        offset = 3 if (len(final_data[middle_name][index].triplets)>1) else 0

        # adds the number of rows required to be skipped
        x_cord += 3*len(final_data[middle_name][index].triplets)+offset
        
    return spaces




# writes both schlussel and zentralcodes
def write_right(name, sheet, middle_name, C_replaced_by, final_data, viol_C, n, dict_newC_J):
    # the required data is fetched from the dictionary
    index_list = final_data[name][0] # if index_list[j] replaced -> grey
    data_list = final_data[name][1]

    # starting coordinates
    x_cord = 5
    y_cord = 22

    replacing_C = set(C_replaced_by.values())
    
    dict_newJ_C = {str(j): su for su, j in dict_newC_J.items()}
    #newC_numbers = dict_newJ_C.keys()

    # writing the data to excel
    for index, number in enumerate(index_list):
        
        # Set the font of a cell to the new font object
        if index in viol_C:
            sheet.cell(row=x_cord-1, column=y_cord).font = fontgrey                
            sheet.cell(row=x_cord-1, column=y_cord).value = str(number) + ' →' + str(C_replaced_by[index] + 1)
        elif index in replacing_C:
            sheet.cell(row=x_cord-1, column=y_cord).font = fontgreen
            sheet.cell(row=x_cord-1, column=y_cord).value = number
        elif index >= n:
            sheet.cell(row=x_cord-1, column=y_cord).font = fontblue
            sheet.cell(row=x_cord-1, column=y_cord).value = number + '   ' + dict_newJ_C[number]                 
        else:
            sheet.cell(row=x_cord-1, column=y_cord).value = number

        for row_index, row in enumerate(data_list[index], x_cord - 1):
            for value_index, value in enumerate(row, 2):
                sheet.cell(row=row_index, column=value_index).value = value
        x_cord+=3


# write middle columns for both sheets
def write_middle(name, sheet, page_break, viol_C, C_replaced_by, final_data, spaces, n):
    # the required data is fetched from the dictionary
    data_list = final_data[name]

    # starting coordinates
    x_cord = 4
    y_cord = 22
    
    replacing_C = set(C_replaced_by.values())

    # writing the data to excel
    for index, middleobject in enumerate(data_list):
        x_cord_end = x_cord+3*len(middleobject.triplets)
        row_indices = [_ for _ in range(x_cord,x_cord_end+1) if (_%57==0 and _!=x_cord_end)]

        if index == page_break or row_indices:
            # print("old coord", x_cord)
            x_cord += 61 - x_cord % 57
            # print("new coord", x_cord)
        elif sheet == sheet_Z and index in spaces["zentralzylinder"]:
            x_cord += 3
        elif sheet == sheet_N and index in spaces["normalzylinder"]:
            x_cord += 3

        for triplet in middleobject.triplets:
            
            for value_index,value in enumerate(triplet.row1):
                if value.strip() != '':
                    if (int(value) - 1) in viol_C:
                        sheet.cell(row=x_cord, column=y_cord+value_index).font = fontgrey12
                        
                    elif (int(value) - 1) in replacing_C:
                        sheet.cell(row=x_cord, column=y_cord+value_index).font = fontgreen12
                        
                    elif (int(value) - 1) >= n:
                        sheet.cell(row=x_cord, column=y_cord+value_index).font = fontblue12
                sheet.cell(row=x_cord, column=y_cord+value_index).value = value
            
            for value_index,value in enumerate(triplet.row2):
                if value.strip() != '':
                    if (int(value) - 1) in viol_C:
                        sheet.cell(row=x_cord+1, column=y_cord+value_index).font = fontgrey12
                        
                    elif (int(value) - 1) in replacing_C:
                        sheet.cell(row=x_cord+1, column=y_cord+value_index).font = fontgreen12
                        
                    elif (int(value) - 1) >= n:
                        sheet.cell(row=x_cord+1, column=y_cord+value_index).font = fontblue12
                #print(f"value_index={value_index}, value={value}, type(value)={type(value)}")
                sheet.cell(row=x_cord+1, column=y_cord+value_index).value = value
            
            for value_index,value in enumerate(triplet.row3):
                if value.strip() != '':
                    if (int(value) - 1) in viol_C:
                        sheet.cell(row=x_cord+2, column=y_cord+value_index).font = fontgrey12
                        
                    elif (int(value) - 1) in replacing_C:
                        sheet.cell(row=x_cord+2, column=y_cord+value_index).font = fontgreen12
                        
                    elif (int(value) - 1) >= n:
                        sheet.cell(row=x_cord+2, column=y_cord+value_index).font = fontblue12
                sheet.cell(row=x_cord+2, column=y_cord+value_index).value = value
                
            x_cord += 3
        
        if len(middleobject.triplets)>1:
            x_cord += 3

# segments one zentralzylinder into required rows of 15 cells each
def segment_middle_zentral(row):
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    offsets = {"A":0,"B":26,"C":52,"D":78,"E":104,"F":130}
    data = ["" for _ in range(135)]

    for value in row:
        value = value.strip("\n")
        if len(value)==1:
            data[letters.index(value)] = value

        else:   
            if value.startswith("A"):
                data[letters.index(value[1])] = value[1]
            elif value:
                data[letters.index(value[1])+offsets[value[0]]] = value

    middleobject = MiddleObject()
    index = 0
    while index<len(data):
        if any(data[index:index+15]):
            index+=15
        else:
            del data[index:index+15]

    while data:
        triplet = Triplet(data[:15],data[15:30],data[30:45])
        if triplet.exists():
            middleobject.add_triplet(triplet)
            data = data[45:]
        else:
            temp = ["" for _ in range(15)]
            if len(middleobject.triplets)==0:
                middleobject.add_triplet(Triplet(temp,temp,temp))
            break
    
    # print(middleobject)
    return middleobject


# segments one normalzylinder into required rows of 15 cells each
def segment_middle_normal(row):
    data = ["" for _ in range(135)]
    middleobject = MiddleObject()

    # creates a unidimensional list of the data
    for value in row.split():
        if row=="":
            temp = ["" for _ in range(15)]
            middleobject.add_triplet(Triplet(temp,temp,temp))
            return middleobject
        
        value = value.strip("\n")
        index = int(value)-1
        while len(data)<=index:
            data.extend(["" for _ in range(135)])

        data[int(value)-1] = str(value)

    # removes any rows (len=15) in the data if it is empty    
    index = 0
    while index<len(data):
        if any(data[index:index+15]):
            index+=15
        else:
            del data[index:index+15]
    
    # segments the data into triplets and adds them to the middleobject
    while data:
        triplet = Triplet(data[:15],data[15:30],data[30:45])
        if triplet.exists():
            middleobject.add_triplet(triplet)
            data = data[45:]
        elif not data:
            if len(middleobject.triplets)==0:
                temp = ["" for _ in range(15)]
                middleobject.add_triplet(Triplet(temp,temp,temp))
            break
        else:
            data=data[45:]

    # print(middleobject)
    # print(len(middleobject.triplets))
    return middleobject


# for converting string to three 15 column rows
def fix_middle(final_data):

    zentral = final_data["zentral_middle"]
    fixed_zentral = []

    for row in zentral:
        row = row.split(" ")
        fixed_zentral.append(segment_middle_zentral(row))

    # making normal values
    normal = final_data["normal_middle"]
    fixed_normal = []

    for row in normal:
        fixed_normal.append(segment_middle_normal(row))

    final_data["zentral_middle"] = fixed_zentral
    final_data["normal_middle"] = fixed_normal


# executing writting
"""
fix_middle()

normal_page_break = group_left("normalzylinder", "normal_middle", final_data)


print("writing normalzylinder")
write_left("normalzylinder", sheet_N, normal_page_break,"normal_middle", viol_R, R_replaced_by)       # R_replaced_by, 

print("writing schlussel")
write_right("schlussel", sheet_S, "normal_middle", C_replaced_by)

print("writing middle columns for normalzylinder")
write_middle("normal_middle", sheet_N, normal_page_break, viol_C, C_replaced_by)

# Save & close the excel.
print("saving file")

wb.save("Result/written_standard_codes.xlsx")
wb.close()
"""
# -----------------------------------------------------------------------------


#----------
def get_new_zyl_from_user_input(n, str_Inz_ex, all_newC) -> list: # for new normalzylinder from user input, n=len(A[0])

    # e.g. str_Inz_ex = 'A4 B4 5   \nA4 5\n   6   C45    dfe  B4 \n 8   7', n=44
    #
    #      returns Inz_ex = [{5, 45, 47}, {5, 45}, {46, 48, 6, 47}, {8, 7}] 
    
    substrings = str_Inz_ex.split('\n')
    substrings = [subst.split(' ') for subst in substrings]
    substrings = [[su for su in subst if su != ''] for subst in substrings]
    substrings = [subst for subst in substrings if subst]
    
    all_substr = set().union(*substrings)
    newC = [s for s in all_substr if not s.isdigit()]
    new_C_range = range(n + 1, n + 1 + len(newC))
    
    all_newC  |= set(newC) 
    
    exiC_dict = {s: int(s) for s in all_substr if s.isdigit()}
    newC_dict = {new_str: new_int for (new_str, new_int) in zip(newC, new_C_range)}
    
    exiC_dict.update(newC_dict)
    exi_new_dict = exiC_dict
    
    Inz_ex = [{exi_new_dict[su] for su in subst} for subst in substrings]

    return Inz_ex, newC_dict# <- used in get_extended_existing_zyl_from_user_input()
    
    
def get_extended_existing_zyl_from_user_input(str_InzC_ex, newC_dict, n, all_newC, dict_newC_J_temporarily) -> dict:
    # zyl 1 additionally opened by new A4, A5 and existing 5, zyl 2 additionally opened by existing 10
    # e.g. str_InzC_ex = '1: A4 A5 5   \n  2:  10'
    #      newC_dict = {'C45': 45, 'dfe': 46, 'B4': 47, 'A4': 48}   from get_new_zyl_from_user_input()[1]
    #      n = 44
    #      returns Inz_ex = [1: {46, 49, 5}, 2: {10}] 
    
    substrings = str_InzC_ex.split('\n')
    substrings = [su for su in substrings if su.strip()]
    substrings = [subst.split(' ') for subst in substrings]
    substrings = [[su for su in subst if su != ''] for subst in substrings]
    
    zyl_numbers = [int(sub[0].strip(':')) for sub in substrings]
    
    zyl_schluss = [[su for su in subst if ':' not in su] for subst in substrings]
    
    all_substr = set().union(*zyl_schluss)
    int_substr = [s for s in all_substr if s.isdigit()]
    nin_substr = [s for s in all_substr if s not in int_substr]
    int_dict = {s: int(s) for s in int_substr}
    
    all_newC  |= set(nin_substr)
    
    # appending new schlüssel-numbers to dict starting from max + 1 of newC_dict.values
    next_schl_numb = max(newC_dict.values()) + 1 if newC_dict.values() else n + 1 # last new schlüssel-number already assigned in newC_dict + 1
    InzC_ex_dict = newC_dict
    for s in nin_substr:
        if s not in InzC_ex_dict.keys():
            InzC_ex_dict[s] = next_schl_numb
            next_schl_numb += 1
    
    dict_newC_J_temporarily.update(InzC_ex_dict) # used in 
    
    InzC_ex_dict.update(int_dict) # here we have all new schl-strings that existed in 
    
    InzC_ex = {z_num: {InzC_ex_dict[su] for su in zyl_schl} for z_num, zyl_schl in zip(zyl_numbers, zyl_schluss)}
    
    return InzC_ex


def get_Inz_ex_InzC_ex(n, str_Inz_ex, str_InzC_ex, all_newC, dict_newC_J_temporarily): 
    Inz_ex, newC_dict = get_new_zyl_from_user_input(n, str_Inz_ex, all_newC)
    InzC_ex = get_extended_existing_zyl_from_user_input(str_InzC_ex, newC_dict, n, all_newC, dict_newC_J_temporarily)
    return Inz_ex, InzC_ex
#----------

#---------- new ----------
def prepend_P_to_unnamed_substrings(substrings):
    modified_substrings = []
    count = 1
    for su in substrings:
        if ':' not in su:
            modified_substrings.append(f'?{count}: {su}')
            count += 1
        else:
            modified_substrings.append(su)
    return modified_substrings
    

def extract_names_and_substrings(str_all_inz, m, n):
    """
    e.g.  
            str_all_inz = ' 1: A4 A5    5 \n  10 v5   \n  ba 4 :  10\n\n 31: B4 7 dfe\n\n  B4 7 66 rrfe\n\n'
    returns ['1', '2', '31'], [['A4', 'A5', '5'], ['10'], ['B4', '7', 'DFE']]
    """
    substrings = str_all_inz.split('\n')
    substrings = [su for su in substrings if su.strip()]
    substrings = [su.upper() for su in substrings]
    substrings = prepend_P_to_unnamed_substrings(substrings)
    names_and_substrings = [sub.split(':') for sub in substrings]
    names_and_substrings = [[s.strip() for s in ns] for ns in names_and_substrings]
    names = [ns[0] for ns in names_and_substrings]
    substrings = [ns[1:] for ns in names_and_substrings]
    substrings = [ns[0].split() for ns in substrings]
    
    all_C = set().union(*substrings)
    exi_C = [s for s in all_C if (s.isdigit() and int(s) <= n) or (s.lstrip('-').isdigit() and abs(int(s)) <= n)]   # int(s) if s.lstrip('-').isdigit() else
    new_C = [s for s in all_C if s not in exi_C] 
    
    all_R = names
    exi_R = [s for s in all_R if s.isdigit() and int(s) <= m]
    new_R = [s for s in all_R if s not in exi_R] 
    
    new_C_range = range(n + 1, n + 1 + len(new_C))
    dict_newC_J_temporarily = {new_c: new_j for (new_c, new_j) in zip(new_C, new_C_range)}
    
    new_R_range = range(m + 1, m + 1 + len(new_R))
    dict_newR_I_temporarily = {new_r: new_i for (new_r, new_i) in zip(new_R, new_R_range)}
    
    dict_Inz_ex_s = dict()
    Inz_ex_s = []
    InzC_ex_s = dict()
    
    for i, name in enumerate(names):
        if name.isdigit():
            if int(name) <= m:
                InzC_ex_s[int(name)] = substrings[i]
            else:
                Inz_ex_s.append(substrings[i])
                dict_Inz_ex_s[name] = substrings[i]
        else:
            Inz_ex_s.append(substrings[i])
            dict_Inz_ex_s[name] = substrings[i]
            
    Inz_ex = [{1} | {dict_newC_J_temporarily[s] if s in new_C else int(s) for s in inz_ex} for inz_ex in Inz_ex_s]
    
    InzC_ex = {i: {dict_newC_J_temporarily[s] if s in new_C else int(s) for s in InzC_ex_s[i]} for i in InzC_ex_s.keys()}
    
    return Inz_ex, InzC_ex, dict_newC_J_temporarily, dict_newR_I_temporarily
#---------- new ----------

def get_newC_name_ind_assignment(C_replaced_by, A_ex, str_Inz_ex, str_InzC_ex, all_newC, m, n, m_ex, n_ex) -> dict: #erase this function
    """
    new zylinders are                                                             new once
                        i = m + len(R_replaced_by), ..., m + len(R_replaced_by) + len(str_Inz_ex)  = m_ex
    .
    new schlüssel are   j = n + len(C_replaced_by), ..., n + len(C_replaced_by) + len(all_newC)    = n_ex
    .
    Out of them, for each new_c in all_newC 
    
    """
    m_start_new_I = m + len(R_replaced_by)
    n_start_new_J = n + len(C_replaced_by)
    new_I = range(m_start_new_I, m_ex)
    new_J = range(n_start_new_J, n_ex)
    inz_new_J_new_I = {j: {i for i in new_I if A_ex[i, j]}}  #for each new j its incidence-set of new i's.
    
    pass

def display_neufertigen(lbl_neufertigen, R_replaced_by, C_replaced_by): 
    
    neuf_R = list(R_replaced_by.keys())
    neuf_C = list(C_replaced_by.keys())
    
    str_neufertigen_C = ''
    str_neufertigen_R = ''
    
    #if (not neuf_R) and neuf_C:
    if len(neuf_C) > 1:
        neuf_C_str = ', '.join(str(key + 1) for key in neuf_C[:-1]) + ' und ' + str(neuf_C[-1] + 1)
        str_neufertigen_C = '- Schlüssel mit Schliessnummer ' + neuf_C_str + ' müssen neu gefertigt werden.'
    elif len(neuf_C) == 1:
        str_neufertigen_C = '- Der Schlüssel mit Schliessnummer ' + str(neuf_C[0] + 1) + ' muss neu gefertigt werden.'
    else:
        str_neufertigen_C = '- Keine Schlüssel müssen neu gefertigt werden'
        if len(neu_R) == 0:
            lbl_neufertigen["text"] = 'Nichts muss ersetzt werden'
            return
            
    if len(neuf_R) > 1:
        neuf_R_str = ', '.join(str(key + 1) for key in neuf_R[:-1]) + ' und ' + str(neuf_R[-1] + 1)
        str_neufertigen_R += '- Zylinder mit Schliessnummer ' + neuf_R_str + ' müssen neu gefertigt werden.'
    elif len(neuf_R) == 1:
        str_neufertigen_R = '- Der Zylinder mit Schliessnummer ' + str(neuf_R[0] + 1) + ' muss neu gefertigt werden.'
    else:
        str_neufertigen_R = '- Keine Zylinder müssen neu gefertigt werden'
    
    lbl_neufertigen["text"] ='NEUFERTIGEN\n\n' +  str_neufertigen_C + '\n' + str_neufertigen_R
    

def display_vorgehen(lbl_vorgehen, R_replaced_by, C_replaced_by, dict_newC_J, dict_newR_I):
    
    # to state repleced index i by i.1 (and j by j.1)
    R_replacing_of = {i_replacing: i_replaced for i_replaced, i_replacing in R_replaced_by.items()}           #go on here--------------,,,fidaiiiiiiii
    C_replacing_of = {j_replacing: j_replaced for j_replaced, j_replacing in C_replaced_by.items()}
    
    # must be appended by increasing order of replacing index
    sorted_R_replacing = sorted(R_replacing_of.keys())
    sorted_C_replacing = sorted(C_replacing_of.keys())
    
    # appending replacing zylinders
    if len(sorted_R_replacing) > 1:
        einfügen_R_str = ',  '.join(str(R_replacing_of[i] + 1) + '.1' for i in sorted_R_replacing[:-1]) + '  und  ' + str(R_replacing_of[sorted_R_replacing[-1]] + 1) + '.1'
        str_vorgehen_repl_R = 'Füge nacheinander die Ersatz-Zylinderpositionen  ' + einfügen_R_str + '  ein.'
    elif len(sorted_R_replacing) == 1:
        str_vorgehen_repl_R = 'Füge die Ersatz-Zylinderposition  ' + str(R_replacing_of[sorted_R_replacing[0]] + 1) + '.1' + '  ein.'
    else:
        str_vorgehen_repl_R = ''
    # appending replacing schlüssel
    if len(sorted_C_replacing) > 1:
        einfügen_C_str = ',  '.join(str(C_replacing_of[j] + 1) + '.1' for j in sorted_C_replacing[:-1]) + '  und  ' + str(C_replacing_of[sorted_C_replacing[-1]] + 1) + '.1'
        str_vorgehen_repl_C = 'Füge nacheinander die Ersatzschlüssel  ' + einfügen_C_str + '  ein.'
    elif len(sorted_C_replacing) == 1:
        str_vorgehen_repl_C = 'Füge den Ersatzschlüssel  ' + str(C_replacing_of[sorted_C_replacing[0]] + 1) + '.1' + '  ein.'
    else:
        str_vorgehen_repl_C = ''
    
    # to state index i by zylinder-name (and j by schlüssel-name)
    dict_newI_R = {i_new: name for name, i_new in dict_newR_I.items()}      
    dict_newJ_C = {j_new: name for name, j_new in dict_newC_J.items()}   
    
    # must be appended by increasing order of new indices
    sorted_R_new = sorted(dict_newI_R.keys())
    sorted_C_new = sorted(dict_newJ_C.keys())        
    
    # appending new zylinders
    if len(sorted_R_new) > 1:
        einfügen_new_R_str = ',  '.join(dict_newI_R[i] for i in sorted_R_new[:-1]) + '  und  ' + dict_newI_R[sorted_R_new[-1]] 
        str_vorgehen_new_R = 'Füge nacheinander die neuen Zylinderpositionen  ' + einfügen_new_R_str + '  ein.'
    elif len(sorted_R_replacing) == 1:
        str_vorgehen_new_R = 'Füge die neue Zylinderposition  ' + dict_newI_R[sorted_R_new[0]] + '  ein.'
    else:
        str_vorgehen_new_R = ''
    # appending new schlüssel
    if len(sorted_C_new) > 1:
        einfügen_new_C_str = ',  '.join(dict_newJ_C[i] for i in sorted_C_new[:-1]) + '  und  ' + dict_newJ_C[sorted_C_new[-1]] 
        str_vorgehen_new_C = 'Füge nacheinander die neuen Schlüssel  ' + einfügen_new_C_str + '  ein.'
    elif len(sorted_C_new) == 1:
        str_vorgehen_new_C = 'Füge den neuen Schlüssel  ' + dict_newJ_C[sorted_C_new[0]] + '  ein.'
    else:
        str_vorgehen_new_C = ''
    
    str_kreuze = 'Kreuze und codiere gemäss \'written_standard_codes.xlsx\'.'    
    #co1 = '1' if len(str_vorgehen_R) else ''
    #co2 = '2' if len(str_vorgehen_R) else ''

    #lbl_vorgehen["text"] = str_vorgehen_repl_R + '\n' + str_vorgehen_new_R + '\n'  + str_vorgehen_repl_C + '\n' + str_vorgehen_new_C + '\n' + str_kreuze

    lines = [str_vorgehen_repl_R, str_vorgehen_new_R, str_vorgehen_repl_C, str_vorgehen_new_C, str_kreuze]
    non_empty_lines = [line for line in lines if line]
    indexed_lines = [f"{i+1}. {line}" for i, line in enumerate(non_empty_lines)]
    lbl_vorgehen["text"] = '\nVORGEHEN\n\n' + "\n".join(indexed_lines)


def solve_extension(str_all_inz, lbl_vorgehen, lbl_neufertigen, mixed=False):#(str_Inz_ex, str_InzC_ex, lbl_vorgehen):
    
    # getting appropriate datas of current R_pr, C_pr, domain D_ext and A
    R_pr, C_pr, R_original_indices, C_original_indices, final_data = get_prev_data()

    resulting_middle_N = [[int(num) for num in string.split()] for string in final_data["normal_middle"]]
    A = binA(resulting_middle_N)
    m, n = A.shape

    D_ext = get_domain_int(R_pr, C_pr, D, final_data, mixed) #this D_ext is domain ready for computing
    
    all_newC = set()
    dict_newC_J_temporarily = dict()
    dict_newR_I_temporarily = dict() 
    
    #Inz_ex, InzC_ex = get_Inz_ex_InzC_ex(n, str_Inz_ex, str_InzC_ex, all_newC, dict_newC_J_temporarily) #### renew.....
    
    Inz_ex, InzC_ex, dict_newC_J_temporarily, dict_newR_I_temporarily = extract_names_and_substrings(str_all_inz, m, n) 
    
    A_ex = get_A_ex(A, Inz_ex, InzC_ex)

    repl = [] #dummy
    R_ex_sol, C_ex_sol, viol_R, viol_C = CSP_ex(A, A_ex, D_ext, R_pr, C_pr, repl, bitlength=26)

    m_ex, n_ex = A_ex.shape


    final_R, final_C, middles, R_replaced_by, C_replaced_by = get_extension_middles(R_pr, C_pr, R_ex_sol, C_ex_sol, viol_R, viol_C)
    
    dict_newC_J = {su: j + len(C_replaced_by) for su, j in dict_newC_J_temporarily.items()}              # final 'new schlüssel : j' dict (j 1-indexed)
    dict_newR_I = {su: i + len(R_replaced_by) for su, i in dict_newR_I_temporarily.items()}
    
    #impose 'Vorgehen' from replaced ....TODO
    display_neufertigen(lbl_neufertigen, R_replaced_by, C_replaced_by)   
    display_vorgehen(lbl_vorgehen, R_replaced_by, C_replaced_by, dict_newC_J, dict_newR_I)        

    final_C = impose_both_7_into_C(final_C, final_data) #get 0 at both 7-pos
      
    final_R_data = transform_to_codes(final_R, final_data["schlussel"][1][0])
    final_C_data = transform_to_codes(final_C, final_data["schlussel"][1][0])

    #updating final_data for writting
    m_ex, n_ex = len(final_R), len(final_C)
    final_m, final_n = len(final_R_data), len(final_C_data)
    final_data["normalzylinder"] = ([str(i + 1) for i in range(final_m)], final_R_data)
    final_data["schlussel"] = ([str(j + 1) for j in range(final_n)], final_C_data)   ###### !!!!!!!!!!!!
    final_data["normal_middle"] = middles

    #print(final_data["normalzylinder"][1])

    colors_data = read_colors()
    #colors_data['normalzylinder'] += colors_data['normalzylinder'][29:]     # temporarily fixed for 33 +  keys      
    append_colors(colors_data, final_data, 'normalzylinder', final_m) 

    """
    wb = openpyxl.load_workbook("Data/raw.xlsx")
    sheet_one = wb[wb.sheetnames[0]]
    sheet_N = wb[wb.sheetnames[1]]
    sheet_S = wb[wb.sheetnames[2]]
    sheet_Z = wb[wb.sheetnames[3]]
    sheet_ZC = wb[wb.sheetnames[4]]
    """

    # refers to page breaks for both sheets, -1 initially
    zentral_page_break = -1
    normal_page_break = -1

    # used for assigning scores to each row, while sorting them, one for red and one for black
    red_scores = {}
    black_scores = {}

    # used for spaces in zentralzylinder and normalzylinder
    spaces = {}
    
    fix_middle(final_data)          

    normal_page_break, spaces = group_left("normalzylinder", "normal_middle", final_data, colors_data, spaces)

    print("writing normalzylinder")
    spaces = write_left("normalzylinder", sheet_N, normal_page_break,"normal_middle", viol_R, R_replaced_by, final_data, spaces, m, dict_newR_I)       # R_replaced_by, 

    print("writing schlussel")
    write_right("schlussel", sheet_S, "normal_middle", C_replaced_by, final_data, viol_C, n, dict_newC_J)

    print("writing middle columns for normalzylinder")
    write_middle("normal_middle", sheet_N, normal_page_break, viol_C, C_replaced_by, final_data, spaces, n)

    # Save & close the excel.
    print("saving file")

    wb.save("Result/written_standard_codes.xlsx")
    wb.close()


if __name__ == '__main__':
    print("extension.py invoked")



























